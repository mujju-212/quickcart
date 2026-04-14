"""
Report Generation Routes for Admin
Provides efficient export functionality in PDF and Excel formats
Uses optimized queries with proper indexing
"""
from flask import Blueprint, jsonify, request, send_file
from datetime import datetime, timedelta
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.pdfgen import canvas
from psycopg2.extras import RealDictCursor
from backend.utils.database import db
from backend.utils.auth_middleware import admin_required
from decimal import Decimal

report_bp = Blueprint('reports', __name__)

# ============= HELPER FUNCTIONS =============

def get_selected_fields(request_fields, all_fields):
    """Parse and validate selected fields from request"""
    if not request_fields:
        return all_fields
    
    requested = request_fields.split(',')
    # Only return fields that exist in all_fields
    return [f for f in requested if f in all_fields]

def is_preview_request():
    """Check if this is a preview request"""
    return request.args.get('preview', 'false').lower() == 'true'

def format_currency(amount):
    """Format amount as currency"""
    return f"₹{float(amount):,.2f}"

def get_date_range(range_type):
    """Get start and end date based on range type"""
    today = datetime.now().date()
    
    if range_type == 'today':
        return today, today
    elif range_type == 'week':
        return today - timedelta(days=7), today
    elif range_type == 'month':
        return today - timedelta(days=30), today
    elif range_type == 'year':
        return today - timedelta(days=365), today
    else:
        return None, None

def create_excel_workbook(title):
    """Create a new Excel workbook with styling"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name limit
    
    # Header styling
    header_fill = PatternFill(start_color='FFE01B', end_color='FFE01B', fill_type='solid')
    header_font = Font(bold=True, size=12)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    return wb, ws, header_fill, header_font, header_alignment

def apply_excel_styling(ws, headers, header_fill, header_font, header_alignment):
    """Apply consistent styling to Excel worksheet"""
    # Set headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

# ============= PRODUCTS REPORT =============

@report_bp.route('/products', methods=['GET'])
@admin_required
def export_products(admin_user):
    """
    Export products report with filters
    Query params: category_id, stock_status, format (pdf/excel), preview (true/false), fields (comma-separated)
    """
    try:
        category_id = request.args.get('category_id', 'all')
        stock_status = request.args.get('stock_status', 'all')  # all, in_stock, out_of_stock, low_stock
        export_format = request.args.get('format', 'excel')
        preview = is_preview_request()
        selected_fields = request.args.get('fields', '')
        
        # Define field mapping
        all_fields = ['product_id', 'name', 'category', 'price', 'stock', 'status', 'description', 'image_url', 'created_at']
        fields_to_select = get_selected_fields(selected_fields, all_fields) if selected_fields else all_fields
        
        with db.get_connection() as conn:
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            # Build optimized query with proper indexing
            query = '''
                SELECT 
                    p.id as product_id,
                    p.name,
                    c.name as category,
                    p.price,
                    p.stock,
                    p.status,
                    p.description,
                    p.image_url,
                    p.created_at
                FROM products p
                LEFT JOIN categories c ON p.category_id = c.id
                WHERE 1=1
            '''
            
            params = []
            
            # Apply filters
            if category_id != 'all':
                query += ' AND p.category_id = %s'
                params.append(int(category_id))
            
            if stock_status == 'in_stock':
                query += ' AND p.stock > 10'
            elif stock_status == 'out_of_stock':
                query += ' AND p.stock = 0'
            elif stock_status == 'low_stock':
                query += ' AND p.stock > 0 AND p.stock <= 10'
            
            query += ' ORDER BY p.id'
            
            cursor.execute(query, params)
            products = cursor.fetchall()
            
            # Filter data by selected fields
            filtered_data = []
            for product in products:
                filtered_row = {field: product.get(field) for field in fields_to_select}
                # Format values
                if 'price' in filtered_row and filtered_row['price']:
                    filtered_row['price'] = f"₹{float(filtered_row['price'])}"
                if 'created_at' in filtered_row and filtered_row['created_at']:
                    filtered_row['created_at'] = filtered_row['created_at'].strftime('%Y-%m-%d')
                filtered_data.append(filtered_row)
            
            # If preview, return JSON
            if preview:
                return jsonify({
                    'success': True,
                    'data': filtered_data,  # Return all data for preview
                    'total': len(filtered_data),
                    'fields': fields_to_select
                })
            
            # Generate report based on format
            if export_format == 'excel':
                return generate_products_excel(filtered_data, fields_to_select)
            else:
                return generate_products_pdf(filtered_data, fields_to_select)
                
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

def generate_products_excel(products, selected_fields):
    """Generate Excel file for products"""
    wb, ws, header_fill, header_font, header_alignment = create_excel_workbook('Products Report')
    
    # Field labels mapping
    field_labels = {
        'product_id': 'ID',
        'name': 'Name',
        'category': 'Category',
        'price': 'Price',
        'stock': 'Stock',
        'status': 'Status',
        'description': 'Description',
        'image_url': 'Image URL',
        'created_at': 'Created Date'
    }
    
    # Headers based on selected fields
    headers = [field_labels.get(f, f) for f in selected_fields]
    
    apply_excel_styling(ws, headers, header_fill, header_font, header_alignment)
    
    # Data rows
    for row_num, product in enumerate(products, 2):
        for col_num, field in enumerate(selected_fields, 1):
            value = product.get(field, 'N/A')
            ws.cell(row=row_num, column=col_num, value=str(value) if value is not None else 'N/A')
    
    # Add summary sheet
    summary_ws = wb.create_sheet('Summary')
    summary_ws['A1'] = 'Total Products'
    summary_ws['B1'] = len(products)
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"products_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

def generate_products_pdf(products, selected_fields):
    """Generate PDF file for products"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=24, textColor=colors.HexColor('#000000'), spaceAfter=30)
    
    # Title
    title = Paragraph("Products Report", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Field labels mapping
    field_labels = {
        'product_id': 'ID',
        'name': 'Name',
        'category': 'Category',
        'price': 'Price',
        'stock': 'Stock',
        'status': 'Status',
        'description': 'Description',
        'image_url': 'Image URL',
        'created_at': 'Created Date'
    }
    
    # Products table (first 50 products for PDF)
    table_data = [[field_labels.get(f, f) for f in selected_fields]]
    for product in products[:50]:
        row = [str(product.get(f, 'N/A'))[:30] for f in selected_fields]  # Truncate long values
        table_data.append(row)
    
    # Calculate column widths dynamically
    num_cols = len(selected_fields)
    col_width = 7.0 / num_cols  # Distribute 7 inches across columns
    col_widths = [col_width * inch] * num_cols
    
    products_table = Table(table_data, colWidths=col_widths)
    products_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FFE01B')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(products_table)
    
    if len(products) > 50:
        note = Paragraph(f"<i>Note: Showing first 50 of {len(products)} products. Download Excel for complete data.</i>", styles['Normal'])
        elements.append(Spacer(1, 0.2*inch))
        elements.append(note)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"products_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )

# ============= ORDERS REPORT =============

@report_bp.route('/orders', methods=['GET'])
@admin_required
def export_orders(admin_user):
    """
    Export orders report with filters
    Query params: status, date_range, date_from, date_to, format, preview (true/false), fields (comma-separated)
    """
    try:
        status = request.args.get('status', 'all')
        date_range = request.args.get('date_range', 'all')
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')
        export_format = request.args.get('format', 'excel')
        preview = is_preview_request()
        selected_fields = request.args.get('fields', '')
        
        # Define field mapping
        all_fields = ['order_id', 'customer_name', 'phone', 'total_amount', 'status', 'order_date', 'address', 'items_count', 'payment_method']
        fields_to_select = get_selected_fields(selected_fields, all_fields) if selected_fields else all_fields
        
        with db.get_connection() as conn:
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            # Optimized query with JOIN
            query = '''
                SELECT 
                    o.id as order_id,
                    o.user_name as customer_name,
                    o.phone,
                    o.total as total_amount,
                    o.status,
                    o.created_at as order_date,
                    o.delivery_address as address,
                    (SELECT COUNT(*) FROM order_items oi WHERE oi.order_id = o.id) as items_count,
                    o.payment_method
                FROM orders o
                WHERE 1=1
            '''
            
            params = []
            
            # Apply status filter
            if status != 'all':
                query += ' AND o.status = %s'
                params.append(status)
            
            # Apply date range filter
            if date_range != 'all' and date_range != 'custom':
                start_date, end_date = get_date_range(date_range)
                if start_date and end_date:
                    query += ' AND o.order_date BETWEEN %s AND %s'
                    params.extend([start_date, end_date])
            elif date_range == 'custom' and date_from and date_to:
                query += ' AND o.order_date BETWEEN %s AND %s'
                params.extend([date_from, date_to])
            
            query += ' ORDER BY o.created_at DESC'
            
            cursor.execute(query, params)
            orders = cursor.fetchall()
            
            # Filter data by selected fields
            filtered_data = []
            for order in orders:
                filtered_row = {field: order.get(field) for field in fields_to_select}
                # Format values
                if 'total_amount' in filtered_row and filtered_row['total_amount']:
                    filtered_row['total_amount'] = f"₹{float(filtered_row['total_amount'])}"
                if 'order_date' in filtered_row and filtered_row['order_date']:
                    filtered_row['order_date'] = filtered_row['order_date'].strftime('%Y-%m-%d %H:%M')
                filtered_data.append(filtered_row)
            
            # If preview, return JSON
            if preview:
                return jsonify({
                    'success': True,
                    'data': filtered_data,  # Return all data for preview
                    'total': len(filtered_data),
                    'fields': fields_to_select
                })
            
            # Generate report
            if export_format == 'excel':
                return generate_orders_excel(filtered_data, fields_to_select)
            else:
                return generate_orders_pdf(filtered_data, fields_to_select)
                
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

def generate_orders_excel(orders, selected_fields):
    """Generate Excel file for orders"""
    wb, ws, header_fill, header_font, header_alignment = create_excel_workbook('Orders Report')
    
    # Field labels mapping
    field_labels = {
        'order_id': 'Order ID',
        'customer_name': 'Customer',
        'phone': 'Phone',
        'total_amount': 'Total',
        'status': 'Status',
        'order_date': 'Date',
        'address': 'Address',
        'items_count': 'Items',
        'payment_method': 'Payment'
    }
    
    headers = [field_labels.get(f, f) for f in selected_fields]
    
    apply_excel_styling(ws, headers, header_fill, header_font, header_alignment)
    
    # Data rows
    for row_num, order in enumerate(orders, 2):
        for col_num, field in enumerate(selected_fields, 1):
            value = order.get(field, 'N/A')
            ws.cell(row=row_num, column=col_num, value=str(value) if value is not None else 'N/A')
    
    # Summary sheet
    summary_ws = wb.create_sheet('Summary')
    summary_ws['A1'] = 'Total Orders'
    summary_ws['B1'] = len(orders)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"orders_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

def generate_orders_pdf(orders, selected_fields):
    """Generate PDF file for orders"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=24, textColor=colors.HexColor('#000000'), spaceAfter=30)
    
    title = Paragraph("Orders Report", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Field labels mapping
    field_labels = {
        'order_id': 'Order ID',
        'customer_name': 'Customer',
        'phone': 'Phone',
        'total_amount': 'Total',
        'status': 'Status',
        'order_date': 'Date',
        'address': 'Address',
        'items_count': 'Items',
        'payment_method': 'Payment'
    }
    
    # Orders table (first 50 for PDF)
    table_data = [[field_labels.get(f, f) for f in selected_fields]]
    for order in orders[:50]:
        row = [str(order.get(f, 'N/A'))[:30] for f in selected_fields]
        table_data.append(row)
    
    # Calculate column widths
    num_cols = len(selected_fields)
    col_width = 7.0 / num_cols
    col_widths = [col_width * inch] * num_cols
    
    orders_table = Table(table_data, colWidths=col_widths)
    orders_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FFE01B')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(orders_table)
    
    if len(orders) > 50:
        note = Paragraph(f"<i>Note: Showing first 50 of {len(orders)} orders. Download Excel for complete data.</i>", styles['Normal'])
        elements.append(Spacer(1, 0.2*inch))
        elements.append(note)
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"orders_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )

# Continue with other report endpoints in next file...
# ============= USERS REPORT =============


@report_bp.route('/users', methods=['GET'])
@admin_required
def export_users(admin_user):
    """
    Export users report with filters
    Query params: role, status, format, preview, fields
    """
    try:
        preview = is_preview_request()
        role = request.args.get('role', 'all')
        status = request.args.get('status', 'all')
        export_format = request.args.get('format', 'excel')
        selected_fields = request.args.get('fields', '')
        
        # Define all available fields
        all_fields = ['user_id', 'name', 'email', 'phone', 'role', 'status', 'created_at', 
                      'last_login', 'login_count', 'total_orders', 'total_spent', 'last_order_date']
        fields_to_select = get_selected_fields(selected_fields, all_fields)
        
        with db.get_connection() as conn:
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            query = '''
                SELECT 
                    u.id as user_id,
                    u.name,
                    u.phone,
                    u.email,
                    u.role,
                    u.status,
                    u.created_at,
                    u.last_login,
                    u.login_count,
                    COALESCE(order_stats.total_orders, 0) as total_orders,
                    COALESCE(order_stats.total_spent, 0) as total_spent,
                    order_stats.last_order_date
                FROM users u
                LEFT JOIN (
                    SELECT 
                        user_id,
                        COUNT(*) as total_orders,
                        SUM(total) as total_spent,
                        MAX(created_at) as last_order_date
                    FROM orders
                    WHERE status != 'cancelled'
                    GROUP BY user_id
                ) order_stats ON u.id = order_stats.user_id
                WHERE 1=1
            '''
            
            params = []
            
            if role != 'all':
                query += ' AND u.role = %s'
                params.append(role)
            
            if status != 'all':
                query += ' AND u.status = %s'
                params.append(status)
            
            query += ' ORDER BY u.created_at DESC'
            
            cursor.execute(query, params)
            users_raw = cursor.fetchall()
            
            # Filter by selected fields
            filtered_data = []
            for user in users_raw:
                filtered_user = {}
                for field in fields_to_select:
                    value = user.get(field)
                    # Format datetime fields
                    if field in ['created_at', 'last_login', 'last_order_date'] and value:
                        filtered_user[field] = value.strftime('%Y-%m-%d %H:%M:%S') if hasattr(value, 'strftime') else str(value)
                    else:
                        filtered_user[field] = value
                filtered_data.append(filtered_user)
            
            if preview:
                return jsonify({
                    'success': True,
                    'data': filtered_data,  # Return all data for preview
                    'total': len(filtered_data)
                })
            
            if export_format == 'excel':
                return generate_users_excel(users_raw, fields_to_select)
            else:
                return generate_users_pdf(users_raw, fields_to_select)
                
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

def generate_users_excel(users, selected_fields=None):
    """Generate Excel for users"""
    if selected_fields is None:
        selected_fields = ['user_id', 'name', 'email', 'phone', 'role', 'status', 'created_at', 
                          'last_login', 'login_count', 'total_orders', 'total_spent', 'last_order_date']
    
    wb, ws, header_fill, header_font, header_alignment = create_excel_workbook('Users Report')
    
    headers = ['ID', 'Name', 'Phone', 'Email', 'Role', 'Status', 'Registration Date', 
               'Last Login', 'Login Count', 'Total Orders', 'Total Spent', 'Last Order Date']
    
    apply_excel_styling(ws, headers, header_fill, header_font, header_alignment)
    
    for row_num, user in enumerate(users, 2):
        ws.cell(row=row_num, column=1, value=user['id'])
        ws.cell(row=row_num, column=2, value=user['name'])
        ws.cell(row=row_num, column=3, value=user['phone'])
        ws.cell(row=row_num, column=4, value=user['email'] or 'N/A')
        ws.cell(row=row_num, column=5, value=user['role'])
        ws.cell(row=row_num, column=6, value=user['status'])
        ws.cell(row=row_num, column=7, value=user['created_at'].strftime('%Y-%m-%d') if user['created_at'] else 'N/A')
        ws.cell(row=row_num, column=8, value=user['last_login'].strftime('%Y-%m-%d %H:%M') if user['last_login'] else 'Never')
        ws.cell(row=row_num, column=9, value=user['login_count'])
        ws.cell(row=row_num, column=10, value=user['total_orders'])
        ws.cell(row=row_num, column=11, value=f"₹{float(user['total_spent'])}")
        ws.cell(row=row_num, column=12, value=user['last_order_date'].strftime('%Y-%m-%d') if user['last_order_date'] else 'N/A')
    
    # Summary
    summary_ws = wb.create_sheet('Summary')
    summary_ws['A1'] = 'Total Users'
    summary_ws['B1'] = len(users)
    summary_ws['A2'] = 'Active Users'
    summary_ws['B2'] = sum(1 for u in users if u['status'] == 'active')
    summary_ws['A3'] = 'Total Revenue'
    total_revenue = sum(float(u['total_spent']) for u in users)
    summary_ws['B3'] = f"₹{total_revenue:,.2f}"
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"users_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

def generate_users_pdf(users):
    """Generate PDF for users"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=24, textColor=colors.HexColor('#000000'), spaceAfter=30)
    
    title = Paragraph("Users Report", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Summary
    total_revenue = sum(float(u['total_spent']) for u in users)
    active_users = sum(1 for u in users if u['status'] == 'active')
    
    summary_data = [
        ['Total Users:', str(len(users))],
        ['Active Users:', str(active_users)],
        ['Total Revenue:', format_currency(total_revenue)],
    ]
    summary_table = Table(summary_data, colWidths=[2*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#FFE01B')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Users table
    table_data = [['Name', 'Phone', 'Role', 'Orders', 'Spent']]
    for user in users[:40]:
        table_data.append([
            user['name'][:25],
            user['phone'],
            user['role'],
            str(user['total_orders']),
            format_currency(user['total_spent'])
        ])
    
    users_table = Table(table_data, colWidths=[2.2*inch, 1.5*inch, 1*inch, 0.8*inch, 1.5*inch])
    users_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FFE01B')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(users_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"users_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )

# ============= INVENTORY REPORT =============

@report_bp.route('/inventory', methods=['GET'])
@admin_required
def export_inventory(admin_user):
    """Export inventory report with preview support"""
    try:
        preview = is_preview_request()
        export_format = request.args.get('format', 'excel')
        selected_fields = request.args.get('fields', '')
        
        # Define all available fields
        all_fields = ['product_id', 'name', 'category', 'price', 'stock', 'stock_value', 'status']
        fields_to_select = get_selected_fields(selected_fields, all_fields)
        
        with db.get_connection() as conn:
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            query = '''
                SELECT 
                    p.id as product_id,
                    p.name,
                    c.name as category,
                    p.price,
                    p.stock,
                    (p.price * p.stock) as stock_value,
                    CASE 
                        WHEN p.stock = 0 THEN 'Out of Stock'
                        WHEN p.stock <= 10 THEN 'Low Stock'
                        WHEN p.stock > 100 THEN 'Overstocked'
                        ELSE 'In Stock'
                    END as status
                FROM products p
                LEFT JOIN categories c ON p.category_id = c.id
                ORDER BY p.stock ASC
            '''
            
            cursor.execute(query)
            products_raw = cursor.fetchall()
            
            # Filter by selected fields
            filtered_data = []
            for product in products_raw:
                filtered_product = {}
                for field in fields_to_select:
                    value = product.get(field)
                    # Format decimal/numeric fields
                    if field in ['price', 'stock_value'] and value:
                        filtered_product[field] = float(value)
                    else:
                        filtered_product[field] = value
                filtered_data.append(filtered_product)
            
            if preview:
                return jsonify({
                    'success': True,
                    'data': filtered_data,  # Return all data for preview
                    'total': len(filtered_data)
                })
            
            if export_format == 'excel':
                return generate_inventory_excel(products_raw, fields_to_select)
            else:
                return generate_inventory_pdf(products_raw, fields_to_select)
                
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

def generate_inventory_excel(products, selected_fields=None):
    """Generate Excel for inventory"""
    wb, ws, header_fill, header_font, header_alignment = create_excel_workbook('Inventory Report')
    
    headers = ['ID', 'Product Name', 'Category', 'Price', 'Stock', 'Stock Value', 'Status']
    apply_excel_styling(ws, headers, header_fill, header_font, header_alignment)
    
    for row_num, product in enumerate(products, 2):
        ws.cell(row=row_num, column=1, value=product['id'])
        ws.cell(row=row_num, column=2, value=product['name'])
        ws.cell(row=row_num, column=3, value=product['category'] or 'N/A')
        ws.cell(row=row_num, column=4, value=f"₹{float(product['price'])}")
        ws.cell(row=row_num, column=5, value=product['stock'])
        ws.cell(row=row_num, column=6, value=f"₹{float(product['stock_value'])}")
        ws.cell(row=row_num, column=7, value=product['stock_status'])
    
    # Summary
    summary_ws = wb.create_sheet('Summary')
    summary_ws['A1'] = 'Total Products'
    summary_ws['B1'] = len(products)
    summary_ws['A2'] = 'Total Stock Value'
    total_value = sum(float(p['stock_value']) for p in products)
    summary_ws['B2'] = f"₹{total_value:,.2f}"
    summary_ws['A3'] = 'Out of Stock'
    summary_ws['B3'] = sum(1 for p in products if p['stock'] == 0)
    summary_ws['A4'] = 'Low Stock'
    summary_ws['B4'] = sum(1 for p in products if 0 < p['stock'] <= 10)
    summary_ws['A5'] = 'In Stock'
    summary_ws['B5'] = sum(1 for p in products if 10 < p['stock'] <= 100)
    summary_ws['A6'] = 'Overstocked'
    summary_ws['B6'] = sum(1 for p in products if p['stock'] > 100)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"inventory_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

def generate_inventory_pdf(products):
    """Generate PDF for inventory"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=24, textColor=colors.HexColor('#000000'), spaceAfter=30)
    
    title = Paragraph("Inventory Report", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Summary
    total_value = sum(float(p['stock_value']) for p in products)
    out_of_stock = sum(1 for p in products if p['stock'] == 0)
    low_stock = sum(1 for p in products if 0 < p['stock'] <= 10)
    
    summary_data = [
        ['Total Products:', str(len(products))],
        ['Total Stock Value:', format_currency(total_value)],
        ['Out of Stock:', str(out_of_stock)],
        ['Low Stock:', str(low_stock)],
    ]
    summary_table = Table(summary_data, colWidths=[2*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#FFE01B')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Low stock products
    low_stock_products = [p for p in products if 0 <= p['stock'] <= 10]
    if low_stock_products:
        heading = Paragraph("<b>Low Stock Alert (≤10 items)</b>", styles['Heading2'])
        elements.append(heading)
        elements.append(Spacer(1, 0.1*inch))
        
        table_data = [['Product', 'Category', 'Stock', 'Value']]
        for product in low_stock_products[:20]:
            table_data.append([
                product['name'][:35],
                (product['category'] or 'N/A')[:20],
                str(product['stock']),
                format_currency(product['stock_value'])
            ])
        
        inventory_table = Table(table_data, colWidths=[2.5*inch, 1.8*inch, 0.8*inch, 1.2*inch])
        inventory_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FFE01B')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(inventory_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"inventory_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )

# ============= CATEGORIES REPORT =============

@report_bp.route('/categories', methods=['GET'])
@admin_required
def export_categories(admin_user):
    """Export categories report with preview support"""
    try:
        preview = is_preview_request()
        export_format = request.args.get('format', 'excel')
        selected_fields = request.args.get('fields', '')
        
        # Define all available fields
        all_fields = ['category_id', 'name', 'products_count', 'status', 'total_revenue', 'total_orders']
        fields_to_select = get_selected_fields(selected_fields, all_fields)
        
        with db.get_connection() as conn:
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            query = '''
                SELECT 
                    c.id as category_id,
                    c.name,
                    COALESCE(prod_count.products_count, 0) as products_count,
                    c.status,
                    COALESCE(cat_revenue.total_revenue, 0) as total_revenue,
                    COALESCE(cat_revenue.total_orders, 0) as total_orders
                FROM categories c
                LEFT JOIN (
                    SELECT 
                        category_id,
                        COUNT(*) as products_count
                    FROM products
                    GROUP BY category_id
                ) prod_count ON c.id = prod_count.category_id
                LEFT JOIN (
                    SELECT 
                        p.category_id,
                        SUM(oi.total_price) as total_revenue,
                        COUNT(DISTINCT oi.order_id) as total_orders
                    FROM order_items oi
                    JOIN products p ON oi.product_id = p.id
                    JOIN orders o ON oi.order_id = o.id
                    WHERE o.status != 'cancelled'
                    GROUP BY p.category_id
                ) cat_revenue ON c.id = cat_revenue.category_id
                ORDER BY c.name
            '''
            
            cursor.execute(query)
            categories_raw = cursor.fetchall()
            
            # Filter by selected fields
            filtered_data = []
            for category in categories_raw:
                filtered_category = {}
                for field in fields_to_select:
                    value = category.get(field)
                    # Handle numeric fields - convert None to 0
                    if field in ['products_count', 'total_orders']:
                        filtered_category[field] = int(value) if value is not None else 0
                    elif field in ['total_revenue']:
                        filtered_category[field] = float(value) if value is not None else 0.0
                    else:
                        filtered_category[field] = value
                filtered_data.append(filtered_category)
            
            if preview:
                return jsonify({
                    'success': True,
                    'data': filtered_data,  # Return all data for preview
                    'total': len(filtered_data)
                })
            
            if export_format == 'excel':
                return generate_categories_excel(categories_raw, fields_to_select)
            else:
                return generate_categories_pdf(categories_raw, fields_to_select)
                
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

def generate_categories_excel(categories, selected_fields=None):
    """Generate Excel for categories"""
    wb, ws, header_fill, header_font, header_alignment = create_excel_workbook('Categories Report')
    
    headers = ['ID', 'Category Name', 'Product Count', 'Status', 'Total Revenue', 'Total Orders']
    apply_excel_styling(ws, headers, header_fill, header_font, header_alignment)
    
    for row_num, category in enumerate(categories, 2):
        ws.cell(row=row_num, column=1, value=category['id'])
        ws.cell(row=row_num, column=2, value=category['name'])
        ws.cell(row=row_num, column=3, value=category['products_count'])
        ws.cell(row=row_num, column=4, value=category['status'])
        ws.cell(row=row_num, column=5, value=f"₹{float(category['total_revenue'])}")
        ws.cell(row=row_num, column=6, value=category['total_orders'])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"categories_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

def generate_categories_pdf(categories):
    """Generate PDF for categories"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=24, textColor=colors.HexColor('#000000'), spaceAfter=30)
    
    title = Paragraph("Categories Report", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Categories table
    table_data = [['Category', 'Products', 'Revenue', 'Orders']]
    for category in categories:
        table_data.append([
            category['name'],
            str(category['products_count']),
            format_currency(category['total_revenue']),
            str(category['total_orders'])
        ])
    
    categories_table = Table(table_data, colWidths=[2.5*inch, 1*inch, 1.5*inch, 1*inch])
    categories_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FFE01B')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(categories_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"categories_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )

# ============= OFFERS REPORT =============

@report_bp.route('/offers', methods=['GET'])
@admin_required  
def export_offers(admin_user):
    """Export offers report with preview support"""
    try:
        preview = is_preview_request()
        export_format = request.args.get('format', 'excel')
        selected_fields = request.args.get('fields', '')
        
        # Define all available fields
        all_fields = ['offer_id', 'title', 'code', 'discount_type', 'discount_value', 'min_order_value',
                      'start_date', 'end_date', 'usage_limit', 'used_count', 'status', 'offer_type']
        fields_to_select = get_selected_fields(selected_fields, all_fields)
        
        with db.get_connection() as conn:
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            query = '''
                SELECT 
                    id as offer_id,
                    title,
                    code,
                    discount_type,
                    discount_value,
                    min_order_value,
                    start_date,
                    end_date,
                    usage_limit,
                    used_count,
                    status,
                    applicable_categories,
                    offer_type
                FROM offers
                ORDER BY start_date DESC
            '''
            
            cursor.execute(query)
            offers_raw = cursor.fetchall()
            
            # Filter by selected fields
            filtered_data = []
            for offer in offers_raw:
                filtered_offer = {}
                for field in fields_to_select:
                    value = offer.get(field)
                    # Format datetime fields
                    if field in ['start_date', 'end_date'] and value:
                        filtered_offer[field] = value.strftime('%Y-%m-%d') if hasattr(value, 'strftime') else str(value)
                    # Format numeric fields
                    elif field in ['discount_value', 'min_order_value'] and value:
                        filtered_offer[field] = float(value)
                    else:
                        filtered_offer[field] = value
                filtered_data.append(filtered_offer)
            
            if preview:
                return jsonify({
                    'success': True,
                    'data': filtered_data,  # Return all data for preview
                    'total': len(filtered_data)
                })
            
            if export_format == 'excel':
                return generate_offers_excel(offers_raw, fields_to_select)
            else:
                return generate_offers_pdf(offers_raw, fields_to_select)
                
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

def generate_offers_excel(offers, selected_fields=None):
    """Generate Excel for offers"""
    wb, ws, header_fill, header_font, header_alignment = create_excel_workbook('Offers Report')
    
    headers = ['ID', 'Title', 'Code', 'Type', 'Discount Value', 'Min Order', 
               'Start Date', 'End Date', 'Usage Limit', 'Used Count', 'Status', 'Offer Type']
    apply_excel_styling(ws, headers, header_fill, header_font, header_alignment)
    
    for row_num, offer in enumerate(offers, 2):
        ws.cell(row=row_num, column=1, value=offer['id'])
        ws.cell(row=row_num, column=2, value=offer['title'])
        ws.cell(row=row_num, column=3, value=offer['code'])
        ws.cell(row=row_num, column=4, value=offer['discount_type'])
        ws.cell(row=row_num, column=5, value=f"₹{float(offer['discount_value'])}" if offer['discount_type'] == 'fixed' else f"{float(offer['discount_value'])}%")
        ws.cell(row=row_num, column=6, value=f"₹{float(offer['min_order_value'])}")
        ws.cell(row=row_num, column=7, value=offer['start_date'].strftime('%Y-%m-%d') if offer['start_date'] else 'N/A')
        ws.cell(row=row_num, column=8, value=offer['end_date'].strftime('%Y-%m-%d') if offer['end_date'] else 'N/A')
        ws.cell(row=row_num, column=9, value=offer['usage_limit'])
        ws.cell(row=row_num, column=10, value=offer['used_count'])
        ws.cell(row=row_num, column=11, value=offer['status'])
        ws.cell(row=row_num, column=12, value=offer['offer_type'])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"offers_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

def generate_offers_pdf(offers):
    """Generate PDF for offers"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=24, textColor=colors.HexColor('#000000'), spaceAfter=30)
    
    title = Paragraph("Offers Report", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Offers table
    table_data = [['Title', 'Code', 'Discount', 'Used', 'Status']]
    for offer in offers:
        discount_val = f"₹{float(offer['discount_value'])}" if offer['discount_type'] == 'fixed' else f"{float(offer['discount_value'])}%"
        table_data.append([
            offer['title'][:30],
            offer['code'],
            discount_val,
            f"{offer['used_count']}/{offer['usage_limit']}",
            offer['status']
        ])
    
    offers_table = Table(table_data, colWidths=[2.2*inch, 1.2*inch, 1*inch, 1*inch, 1*inch])
    offers_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FFE01B')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(offers_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"offers_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )
